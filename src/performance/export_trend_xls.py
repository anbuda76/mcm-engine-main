import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from io import BytesIO


def add_kpi_sheet(wb, name, kpi_per_periodo):
    ws = wb.create_sheet(name)

    rows = []

    for periodo, info in kpi_per_periodo.items():
        if not info["valid"]:
            continue

        df_pen = info["Penetrazione"]
        df_ric = info["Ricordo"]
        df_ing = info["Ingaggio"]

        rows.append({
            "Periodo": periodo,
            "Penetrazione media (%)": round(df_pen["Penetrazione (%)"].mean(), 2),
            "Ricordo medio (%)": round(df_ric["Ricorda (%)"].mean(), 2),
            "Ingaggio medio (%)": round(df_ing["Ingaggio (%)"].mean(), 2),
            "Propensione media": info["Propensione"]["Media propensione"],
            "NPS": info["NPS"]["NPS"]
        })

    df = pd.DataFrame(rows)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)


def add_series_sheet(wb, name, df_series):
    ws = wb.create_sheet(name)

    for r in dataframe_to_rows(df_series, index=False, header=True):
        ws.append(r)

    if df_series.empty:
        return

    chart = LineChart()
    chart.title = name
    chart.y_axis.title = "Valore KPI"

    data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)

    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.set_categories(cats)

    chart.width = 24
    chart.height = 12

    ws.add_chart(chart, "F2")


def export_trend_xlsx(results):
    wb = Workbook()

    # Rimuovo il default sheet
    wb.remove(wb.active)

    # Azienda
    add_kpi_sheet(wb, "KPI_Periodo_AZIENDA", results["Azienda"]["KPI_per_periodo"])
    add_series_sheet(wb, "Serie_Storica_AZIENDA", results["Azienda"]["Series"])

    # Prodotti
    for prod, res in results["Prodotti"].items():
        add_kpi_sheet(wb, f"KPI_Periodo_{prod}", res["KPI_per_periodo"])
        add_series_sheet(wb, f"Serie_Storica_{prod}", res["Series"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer